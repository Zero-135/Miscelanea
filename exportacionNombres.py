import os
import subprocess
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =========================
# CONFIGURACIÓN
# =========================
CARPETA_BASE = r"G:\3-gatsu no Lion\1.2.-3-gatsu no Lion\A"   # ← cambia la ruta
ARCHIVO_EXCEL = r"D:\PycharmProjects\Miscelanea\archivos-titles.xlsx"
EXTENSIONES_VIDEO = (".mp4", ".mkv", ".avi", ".mov", ".webm")

# =========================
# FUNCIONES
# =========================
def obtener_title(video):
    try:
        comando = [
            "ffprobe",
            "-v", "quiet",
            "-print_format", "json",
            "-show_entries", "format_tags=title",
            video
        ]

        resultado = subprocess.run(
            comando,
            capture_output=True,
            text=True,
            encoding="utf-8"
        )

        data = json.loads(resultado.stdout)
        return data.get("format", {}).get("tags", {}).get("title", "")
    except Exception:
        return ""

# =========================
# CREAR EXCEL
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Videos"

# Encabezados
ws.append(["Ruta", "Archivo", "Title"])

# =========================
# PROCESO RECURSIVO
# =========================
for root, _, files in os.walk(CARPETA_BASE):
    for file in files:
        if file.lower().endswith(EXTENSIONES_VIDEO):
            ruta = os.path.join(root, file)
            title = obtener_title(ruta)

            ws.append([
                ruta,
                file,
                title if title else ""
            ])

# =========================
# AJUSTAR COLUMNAS
# =========================
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

# =========================
# ACTIVAR FILTROS
# =========================
ws.auto_filter.ref = ws.dimensions

# =========================
# GUARDAR EXCEL
# =========================
wb.save(ARCHIVO_EXCEL)

# =========================
# ABRIR EXCEL AUTOMÁTICAMENTE
# =========================
os.startfile(ARCHIVO_EXCEL)

print("Proceso finalizado correctamente.")
