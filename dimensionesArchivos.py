import os
from moviepy.editor import VideoFileClip
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ===== CONFIGURACIÓN =====
CARPETA_BASE = r"C:\Users\Walter Rivas\Documents\Videos\Konosuba\Nueva carpeta"   # Ruta donde buscar videos
SALIDA_EXCEL = r"D:\PycharmProjects\Miscelanea\dimensiones_videos.xlsx"


def obtener_dimensiones_video(ruta_video):
    """Retorna (ancho, alto) desde un archivo de video."""
    try:
        clip = VideoFileClip(ruta_video)
        width, height = clip.size
        clip.close()
        return width, height
    except Exception as e:
        print(f"Error leyendo {ruta_video}: {e}")
        return None, None


def escanear_videos(carpeta):
    """Recorre recursivamente y retorna lista con info de videos."""
    datos = []

    for raiz, dirs, archivos in os.walk(carpeta):
        for archivo in archivos:
            if archivo.lower().endswith((".mp4", ".mkv")):
                ruta = os.path.join(raiz, archivo)
                print(f"Procesando: {ruta}")

                w, h = obtener_dimensiones_video(ruta)
                datos.append([ruta, w, h])

    return datos


def auto_ajustar_columnas(ws):
    """Ajusta automáticamente el ancho de las columnas según su contenido."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def exportar_excel(datos, ruta_salida):
    """Exporta la info a un archivo Excel con filtros y autoajuste."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dimensiones"

    # Encabezados
    encabezados = ["Ruta del Video", "Ancho", "Alto"]
    ws.append(encabezados)

    # Datos
    for fila in datos:
        ws.append(fila)

    # Activar filtros
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Auto ajustar columnas
    auto_ajustar_columnas(ws)

    # Guardar Excel
    wb.save(ruta_salida)
    print(f"\nExcel generado: {ruta_salida}")

    # === ABRIR ARCHIVO AL FINAL ===
    try:
        os.startfile(ruta_salida)  # Solo Windows
    except Exception as e:
        print(f"No se pudo abrir el archivo automáticamente: {e}")


# ===== EJECUCIÓN =====
if __name__ == "__main__":
    datos_videos = escanear_videos(CARPETA_BASE)
    exportar_excel(datos_videos, SALIDA_EXCEL)

    # Final del proceso
    print("\nProceso finalizado.")
